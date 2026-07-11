"""Parse participant names from Excel/ODS files."""

from __future__ import annotations

import re
from pathlib import Path

from .common import ResultRow, detect_status, make_row

NAME_HEADERS = {"namn", "name", "deltagare", "runner", "löpare"}
CLUB_HEADERS = {"klubb", "club", "organisation", "org", "förening"}
CLASS_HEADERS = {"klass", "class", "kategori", "bana"}
PLACE_HEADERS = {"plac", "pl", "placering", "rank", "place", "pos"}
TIME_HEADERS = {"tid", "time", "resultat", "löptid"}


def parse_excel_file(path: Path, event_id: int) -> list[ResultRow]:
    suffix = path.suffix.lower()
    if suffix == ".ods":
        return _parse_ods(path, event_id)
    if suffix in {".xlsx", ".xlsm"}:
        return _parse_xlsx(path, event_id)
    if suffix == ".xls":
        return _parse_xls(path, event_id)
    return []


def _parse_xls(path: Path, event_id: int) -> list[ResultRow]:
    import xlrd

    book = xlrd.open_workbook(path)
    rows: list[ResultRow] = []
    for sheet in book.sheets():
        rows.extend(_parse_sheet([sheet.row_values(i) for i in range(sheet.nrows)], event_id))
    return rows


def _parse_xlsx(path: Path, event_id: int) -> list[ResultRow]:
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    rows: list[ResultRow] = []
    for sheet in workbook.worksheets:
        sheet_rows = [[cell.value for cell in row] for row in sheet.iter_rows()]
        rows.extend(_parse_sheet(sheet_rows, event_id))
    return rows


def _parse_ods(path: Path, event_id: int) -> list[ResultRow]:
    from odf.opendocument import load
    from odf.table import Table, TableCell, TableRow
    from odf.text import P

    doc = load(str(path))
    rows: list[ResultRow] = []
    for table in doc.getElementsByType(Table):
        sheet_rows: list[list] = []
        for row in table.getElementsByType(TableRow):
            values = []
            for cell in row.getElementsByType(TableCell):
                text = "".join(
                    node.data for node in cell.getElementsByType(P) if node.firstChild
                )
                values.append(text)
            sheet_rows.append(values)
        rows.extend(_parse_sheet(sheet_rows, event_id))
    return rows


def _parse_sheet(sheet_rows: list[list], event_id: int) -> list[ResultRow]:
    if _is_sylvester_stats_sheet(sheet_rows):
        return []

    rows: list[ResultRow] = []
    header_map: dict[str, int] = {}
    current_class: str | None = None

    for raw_row in sheet_rows:
        cells = [_cell_str(value) for value in raw_row]
        if not any(cells):
            continue

        if not header_map and _looks_like_header(cells):
            header_map = _map_headers(cells)
            continue

        joined = " ".join(cells)
        if len(cells) == 1 and cells[0] and not cells[0][0].isdigit():
            current_class = cells[0]
            continue

        name_idx = header_map.get("name")
        if name_idx is None:
            name_idx = _guess_name_index(cells)
        if name_idx is None or name_idx >= len(cells):
            continue

        name = cells[name_idx]
        status = detect_status(joined)
        place = _parse_int(cells[header_map["place"]]) if "place" in header_map else _parse_leading_place(cells[0])
        time_text = cells[header_map["time"]] if "time" in header_map else _guess_time(cells)
        club = cells[header_map["club"]] if "club" in header_map else None
        class_name = cells[header_map["class"]] if "class" in header_map else current_class

        row = make_row(
            event_id,
            name,
            club=club,
            class_name=class_name,
            place=place,
            time=None if status else time_text,
            status=status,
            parse_source="excel",
            parse_confidence="high",
        )
        if row:
            rows.append(row)

    return rows


def _cell_str(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _looks_like_header(cells: list[str]) -> bool:
    lowered = [cell.lower() for cell in cells]
    return any(header in lowered for headers in (NAME_HEADERS, PLACE_HEADERS, TIME_HEADERS) for header in headers)


def _map_headers(cells: list[str]) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for index, cell in enumerate(cells):
        lowered = cell.lower()
        if lowered in NAME_HEADERS:
            mapping["name"] = index
        elif lowered in CLUB_HEADERS:
            mapping["club"] = index
        elif lowered in CLASS_HEADERS:
            mapping["class"] = index
        elif lowered in PLACE_HEADERS:
            mapping["place"] = index
        elif lowered in TIME_HEADERS:
            mapping["time"] = index
    return mapping


def _guess_name_index(cells: list[str]) -> int | None:
    for index, cell in enumerate(cells):
        if re.search(r"[A-Za-zÅÄÖåäö]{2,}", cell) and not re.fullmatch(r"[\d:,\.]+", cell):
            if index == 0 and _parse_leading_place(cell) is not None:
                continue
            return index
    return None


def _parse_int(value: str | None) -> int | None:
    if not value:
        return None
    match = re.search(r"\d+", value)
    return int(match.group()) if match else None


def _parse_leading_place(value: str) -> int | None:
    match = re.match(r"^(\d+)", value)
    return int(match.group(1)) if match else None


def _guess_time(cells: list[str]) -> str | None:
    for cell in reversed(cells):
        if re.search(r"\d[:,\.]\d", cell):
            return cell.replace(",", ".")
    return None


def _is_sylvester_stats_sheet(sheet_rows: list[list]) -> bool:
    """Detect Sylvesterloppet year-by-year summary sheets, not participant lists."""
    preview = " ".join(_cell_str(value) for row in sheet_rows[:6] for value in row).lower()
    if "sylvesterloppet" not in preview:
        return False

    year_rows = 0
    for raw_row in sheet_rows[2:25]:
        cells = [_cell_str(value) for value in raw_row]
        if not cells or not cells[0]:
            continue
        if re.fullmatch(r"(19|20)\d{2}", cells[0]):
            year_rows += 1

    return year_rows >= 3
